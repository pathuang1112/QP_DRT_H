import sys
import os
import numpy as np
import json
import io
import re
import openpyxl
import pandas as pd
from DRT_1 import DRT_1
from RPT_1 import RPT_1
from ART_1 import ART_1
from RT import  RT
from DRT_H import DRT_H
from DRT_F import DRT_F


if __name__ == '__main__':
    input_qbit_count = int(sys.argv[1])
    output_qbit_count = int(sys.argv[2])
    group_name = sys.argv[3]    #fault version, e.g., qft_8_1
    program_name = sys.argv[4]  #program, e.g. qft_8
    alg = sys.argv[5]   # algorithm, RT RPT_1 ART_1 DRT_1 DRT_F DRT_H
    cand_num = int(sys.argv[6]) # for ART, number of the candicates

    result_path = './result/' + group_name
    pattern1 = re.compile(program_name + '_' + alg + '\.txt$')
    pattern2 = re.compile(program_name + '_' + alg + '_failCount\.txt$')
    pattern3 = re.compile(program_name + '_' + alg + r'\(cand_num=' + str(cand_num) + r'\)\.txt$')
    pattern4 = re.compile(program_name + '_' + alg + r'_failCount\(cand_num=' + str(cand_num) + r'\)\.txt$')

    if os.path.exists(result_path):
        for filename in os.listdir(result_path):
            if pattern1.match(filename) or pattern2.match(filename) or pattern3.match(filename) or pattern4.match(filename) :
                os.remove(os.path.join(result_path, filename))
                print("Delete File: " + os.path.join(result_path, filename))
                print(alg)
    else:
        pass
    ########## Start
    folder = './result/' + group_name
    if not os.path.exists(folder):
        os.makedirs(folder)

    # number of the testcases to be selected
    num_vars = -1  
    if input_qbit_count == 10:
        num_vars = 614 # 2**10*0.3=614.4
    elif input_qbit_count == 11:
        num_vars = 1229  # 2**11*0.3=1228.8
    elif input_qbit_count == 7:
        num_vars = 77  # 2**7*2*0.3=76.8
    elif input_qbit_count == 9:
        num_vars = 307 # 2**9*2*0.3=307.2
    elif input_qbit_count == 8:
        num_vars = 154  # 2**8*2*0.3=153.6
    else:
        num_vars = -1

    iter_num = 50  # number of iterations of each testing strategy
    np.set_printoptions(linewidth=100)  

    if alg == "RT":
        algorithm = RT(num_vars, input_qbit_count, output_qbit_count, group_name, program_name, alg, cand_num,
                        iter_num)  
    elif alg == "ART_1":
        algorithm = ART_1(num_vars, input_qbit_count, output_qbit_count, group_name, program_name, alg, cand_num,
                          iter_num)
    elif alg == "DRT_1":
        algorithm = DRT_1(num_vars, input_qbit_count, output_qbit_count, group_name, program_name, alg, iter_num)
    elif alg == "RPT_1":
        algorithm = RPT_1(num_vars, input_qbit_count, output_qbit_count, group_name, program_name, alg, iter_num)
    elif alg == "DRT_F":
        algorithm = DRT_F(num_vars, input_qbit_count, output_qbit_count, group_name, program_name, alg, iter_num)
    elif alg == "DRT_H":
        algorithm = DRT_H(num_vars, input_qbit_count, output_qbit_count, group_name, program_name, alg, iter_num, os.path.join("./oracle/OpenQASM/" + program_name + "_oracle.xlsx"), -1)
    else:
        print("Search algorithm " + alg + " not supported!")

    algorithm.run()
    
    if alg == "RT":
        best_inputs = algorithm.get_result()
        print(f"RT\'s best inputs:{best_inputs}")
    elif alg == "ART_1":
        best_inputs = algorithm.get_result()
        print(f"ART_1\'s best inputs: {best_inputs}")
    elif alg == "DRT_1":
        best_inputs = algorithm.get_result()
        print(f"DRT_1\'s best inputs: {best_inputs}")
    elif alg == "RPT_1":
        best_inputs = algorithm.get_result()
        print(f"RPT_1\'s best inputs: {best_inputs}")
    elif alg == "DRT_F":
        best_inputs = algorithm.get_result()
        print(f"DRT_F\'s best inputs: {best_inputs}")
    elif alg == "DRT_H":
        best_inputs = algorithm.get_result()
        print(f"DRT_H\'s best inputs: {best_inputs}")
    print(f'Algorithm: ${alg}')
    print(f'FaultVersion: ${group_name}')
    print(f'Computing time: ${algorithm.total_computing_time}')


    ''' Data analysis, output to txt and excel'''
    count_fail = {}
    if alg == 'ART_1' :
        file_path = './result/' + group_name + '/' + program_name + '_' + alg + '(cand_num=' + str(cand_num) + ').txt'
    elif alg == 'RT' or alg == 'DRT_H' or alg == 'DRT_F' or alg == 'DRT_1' or alg == 'RPT_1':
        file_path = './result/' + group_name + '/' + program_name + '_' + alg + '.txt'
    else:
        pass

    with open(file_path, 'r') as f:
        data = f.readlines()
        nums = []
        if alg == 'DRT_1' or alg == 'RPT_1' or alg == 'DRT_F' or alg == 'DRT_H':
            for i in range(3, len(data), 4):
                d = json.load(io.StringIO(data[i].strip()))
                for k, v in d.items():
                    if k not in count_fail:
                        count_fail[k] = [v]
                    else:
                        count_fail[k].append(v)
        else:
            for i in range(2, len(data), 3):
                d = json.load(io.StringIO(data[i].strip()))
                for k, v in d.items():
                    if k not in count_fail:
                        count_fail[k] = [v]
                    else:
                        count_fail[k].append(v)
        print(count_fail)
    avg_failCount = {}
    max_failCount = {}

    for k, v in count_fail.items():
        avg_failCount[k] = sum(v) / len(v)
        max_failCount[k] = max(v)

    print(avg_failCount)
    print(max_failCount)


    with open('./result/' + group_name + '/' + program_name + '_' + alg + '_failCount.txt', "w") as f2:
        f2.write("count_fail:\n")
        f2.write(json.dumps(count_fail) + "\n")
        f2.write("avg_failCount:\n")
        f2.write(json.dumps(avg_failCount) + "\n")
        f2.write("max_failCount:\n")
        f2.write(json.dumps(max_failCount) + "\n")
    if alg == 'ART_1':
        os.rename('./result/' + group_name + '/' + program_name + '_' + alg + '_failCount.txt',
                  './result/' + group_name + '/' + program_name + '_' + alg + '_failCount(cand_num=' + str(cand_num) + ').txt')


    avg_filename = f"{program_name}_avg_faultNum.xlsx"
    max_filename = f"{program_name}_max_faultNum.xlsx"


    try:
        workbook1 = openpyxl.load_workbook('./result/' + avg_filename)
    except FileNotFoundError:
        workbook1 = openpyxl.Workbook()

    if group_name not in workbook1.sheetnames:
        worksheet = workbook1.create_sheet(group_name)
        if alg == 'ART_1' or alg == 'RPT_1' or alg == 'DRT_1' or alg == 'DRT_F' or alg == 'DRT_H':
            worksheet.cell(row=1, column=2, value=alg + "_" + str(cand_num))
        for index, key in enumerate(avg_failCount.keys()):
            worksheet.cell(row=index + 2, column=1, value=key)
        empty_column = 2
    else:
        worksheet = workbook1[group_name]
        empty_column = worksheet.max_column + 1
        if alg == 'ART_1' or alg == 'RPT_1' or alg == 'DRT_1' or alg == 'DRT_F' or alg == 'DRT_H':
            worksheet.cell(row=1, column=empty_column, value=alg + "_" + str(cand_num))

    col_idx = empty_column

    for index, key in enumerate(avg_failCount.keys()):
        worksheet.cell(row=index + 2, column=col_idx, value=avg_failCount[key])
    workbook1.save('./result/' + avg_filename)

    try:
        workbook2 = openpyxl.load_workbook('./result/' + max_filename)
    except FileNotFoundError:
        workbook2 = openpyxl.Workbook()

    if group_name not in workbook2.sheetnames:
        worksheet = workbook2.create_sheet(group_name)
        if alg == 'ART_1' or alg == 'RPT_1' or alg == 'DRT_1' or alg == 'DRT_F' or alg == 'DRT_H':
            worksheet.cell(row=1, column=2, value=alg + "_" + str(cand_num))
        for index, key in enumerate(max_failCount.keys()):
            worksheet.cell(row=index + 2, column=1, value=key)
        empty_column = 2
    else:
        worksheet = workbook2[group_name]
        empty_column = worksheet.max_column + 1
        if alg == 'ART_1' or alg == 'RPT_1' or alg == 'DRT_1' or alg == 'DRT_F' or alg == 'DRT_H':
            worksheet.cell(row=1, column=empty_column, value=alg + "_" + str(cand_num))
    col_idx = empty_column
    for index, key in enumerate(max_failCount.keys()):
        worksheet.cell(row=index + 2, column=col_idx, value=max_failCount[key])
    workbook2.save('./result/' + max_filename)

